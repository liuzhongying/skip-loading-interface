from __future__ import annotations

import json
import mimetypes
import os
import sqlite3
import sys
from io import BytesIO
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
    BUNDLE_DIR = Path(getattr(sys, "_MEIPASS", BASE_DIR))
else:
    BASE_DIR = Path(__file__).resolve().parent
    BUNDLE_DIR = BASE_DIR
DATA_DIR = BASE_DIR / "data"
EXPORT_DIR = BASE_DIR / "exports"
STATIC_DIR = BUNDLE_DIR / "static"
DB_PATH = DATA_DIR / "skip_loading.db"

FAR_FUTURE = date(2099, 12, 31)
DATE_FORMAT = "%Y-%m-%d"
EXCEL_DATE_FORMAT = "dd mmm yyyy"


def ensure_dirs() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    EXPORT_DIR.mkdir(exist_ok=True)


def connect() -> sqlite3.Connection:
    ensure_dirs()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_num TEXT NOT NULL,
                action TEXT NOT NULL CHECK(action IN ('STOP', 'RESUME')),
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                maker TEXT NOT NULL,
                input_date TEXT NOT NULL,
                checker TEXT,
                approval_date TEXT,
                remark TEXT NOT NULL,
                pod TEXT NOT NULL,
                status TEXT NOT NULL CHECK(status IN ('pending', 'approved', 'rejected', 'exported')),
                reject_reason TEXT,
                export_batch_id TEXT,
                created_by TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                previous_request_id INTEGER
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_requests_run_num ON requests(run_num)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_requests_status ON requests(status)"
        )


def today() -> date:
    return date.today()


def parse_date(value: str, field: str) -> date:
    try:
        return datetime.strptime(value, DATE_FORMAT).date()
    except ValueError:
        raise ValueError(f"{field} must be YYYY-MM-DD")


def parse_excel_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d %b %Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Cannot parse Excel date: {text}")


def excel_date_value(value: str | None, field: str) -> date | str:
    if not value:
        return ""
    return parse_date(value, field)


def normalize_run_num(value: str) -> str:
    clean = value.strip()
    if not clean:
        raise ValueError("Run_Num is required")
    return clean


def require_text(payload: dict[str, Any], key: str, label: str) -> str:
    value = str(payload.get(key, "")).strip()
    if not value:
        raise ValueError(f"{label} is required")
    return value


def row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    return {key: row[key] for key in row.keys()}


def latest_effective_request(conn: sqlite3.Connection, run_num: str) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT *
        FROM requests
        WHERE run_num = ?
          AND status IN ('approved', 'exported')
        ORDER BY id DESC
        LIMIT 1
        """,
        (run_num,),
    ).fetchone()


def latest_any_request(conn: sqlite3.Connection, run_num: str) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT *
        FROM requests
        WHERE run_num = ?
        ORDER BY created_at DESC, id DESC
        LIMIT 1
        """,
        (run_num,),
    ).fetchone()


def current_status_from_row(row: sqlite3.Row | None) -> str:
    if row is None:
        return "NO_RECORD"
    if not row["end_date"]:
        return "STOPPED" if row["action"] == "STOP" else "RESUMED"
    end_date = parse_date(row["end_date"], "end_date")
    if row["action"] == "STOP" and end_date >= today():
        return "STOPPED"
    return "RESUMED"


def create_request(payload: dict[str, Any]) -> dict[str, Any]:
    run_num = normalize_run_num(str(payload.get("run_num", "")))
    action = require_text(payload, "action", "Action").upper()
    if action not in {"STOP", "RESUME"}:
        raise ValueError("Action must be STOP or RESUME")

    maker = require_text(payload, "maker", "Maker")
    created_by = str(payload.get("created_by") or maker).strip() or maker
    checker = str(payload.get("checker", "")).strip() or None
    pod = require_text(payload, "pod", "POD")
    remark = require_text(payload, "remark", "Remark")

    now = datetime.now().isoformat(timespec="seconds")
    with connect() as conn:
        previous = latest_any_request(conn, run_num)
        latest_effective = latest_effective_request(conn, run_num)

        if action == "STOP":
            start = today()
            end = FAR_FUTURE
        else:
            if latest_effective is None or latest_effective["action"] != "STOP":
                raise ValueError("RESUME requires an existing current STOP record")
            start = parse_date(latest_effective["start_date"], "Start_Date")
            end = today() - timedelta(days=1)

        cursor = conn.execute(
            """
            INSERT INTO requests (
                run_num, action, start_date, end_date, maker, input_date,
                checker, remark, pod, status, created_by, created_at, updated_at,
                previous_request_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?, ?, ?, ?)
            """,
            (
                run_num,
                action,
                start.isoformat(),
                end.isoformat(),
                maker,
                today().isoformat(),
                checker,
                remark,
                pod,
                created_by,
                now,
                now,
                previous["id"] if previous else None,
            ),
        )
        row = conn.execute(
            "SELECT * FROM requests WHERE id = ?", (cursor.lastrowid,)
        ).fetchone()
        return row_to_dict(row)


def list_requests(
    status: str | None = None,
    run_num: str | None = None,
    checker: str | None = None,
    pod: str | None = None,
) -> list[dict[str, Any]]:
    clauses: list[str] = []
    params: list[Any] = []
    if status:
        clauses.append("status = ?")
        params.append(status)
    if run_num:
        clauses.append("run_num = ?")
        params.append(run_num)
    if checker:
        clauses.append("LOWER(checker) = LOWER(?)")
        params.append(checker.strip())
    if pod:
        clauses.append("LOWER(pod) = LOWER(?)")
        params.append(pod.strip())

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    with connect() as conn:
        rows = conn.execute(
            f"""
            SELECT *
            FROM requests
            {where}
            ORDER BY created_at DESC, id DESC
            LIMIT 500
            """,
            params,
        ).fetchall()
        return [row_to_dict(row) for row in rows]


def get_history(run_num: str) -> dict[str, Any]:
    run_num = normalize_run_num(run_num)
    with connect() as conn:
        history = conn.execute(
            """
            SELECT *
            FROM requests
            WHERE run_num = ?
            ORDER BY created_at DESC, id DESC
            """,
            (run_num,),
        ).fetchall()
        latest = latest_effective_request(conn, run_num)
        return {
            "run_num": run_num,
            "current_status": current_status_from_row(latest),
            "current_record": row_to_dict(latest) if latest else None,
            "history": [row_to_dict(row) for row in history],
        }


def get_defaults(run_num: str | None = None, action: str | None = None) -> dict[str, Any]:
    normalized_action = (action or "STOP").upper()
    if normalized_action not in {"STOP", "RESUME"}:
        raise ValueError("Action must be STOP or RESUME")

    if normalized_action == "STOP":
        return {
            "action": "STOP",
            "start_date": today().isoformat(),
            "end_date": FAR_FUTURE.isoformat(),
            "message": "STOP starts today and stays active until 2099-12-31.",
        }

    run_num_value = normalize_run_num(run_num or "")
    with connect() as conn:
        latest = latest_effective_request(conn, run_num_value)
        if latest is None or latest["action"] != "STOP":
            return {
                "action": "RESUME",
                "start_date": "",
                "end_date": (today() - timedelta(days=1)).isoformat(),
                "message": "RESUME requires an existing current STOP record.",
            }
        return {
            "action": "RESUME",
            "start_date": latest["start_date"],
            "end_date": (today() - timedelta(days=1)).isoformat(),
            "message": "RESUME keeps the current STOP Start_Date and sets End_Date to yesterday.",
        }


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def import_history_excel(file_bytes: bytes) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel is empty")

    headers = {normalize_header(value): index for index, value in enumerate(rows[0])}

    def cell(row: tuple[Any, ...], *names: str) -> Any:
        for name in names:
            index = headers.get(normalize_header(name))
            if index is not None and index < len(row):
                return row[index]
        return ""

    latest_by_run_num: dict[str, dict[str, Any]] = {}
    skipped_blank_run_num = 0
    for row in rows[1:]:
        run_num = str(cell(row, "Run_Num", "Run Num", "Run_Number") or "").strip()
        if not run_num:
            skipped_blank_run_num += 1
            continue

        start_date = parse_excel_date(cell(row, "Start_Date", "Start Date"))
        end_date = parse_excel_date(cell(row, "End_Date", "End Date"))
        input_date = parse_excel_date(cell(row, "Input_Date", "Input Date"))
        approval_date = parse_excel_date(cell(row, "Approval_Date", "Approval Date"))

        action = "RESUME" if end_date is not None and end_date <= today() else "STOP"
        latest_by_run_num[run_num] = {
            "run_num": run_num,
            "action": action,
            "start_date": start_date.isoformat() if start_date else "",
            "end_date": end_date.isoformat() if end_date else "",
            "maker": str(cell(row, "Maker") or "").strip(),
            "input_date": input_date.isoformat() if input_date else "",
            "checker": str(cell(row, "Checker") or "").strip() or None,
            "approval_date": approval_date.isoformat() if approval_date else None,
            "remark": str(cell(row, "REMARK", "Remark") or "").strip(),
            "pod": str(cell(row, "POD") or "").strip(),
        }

    now = datetime.now().isoformat(timespec="seconds")
    imported_ids: list[int] = []
    with connect() as conn:
        for item in latest_by_run_num.values():
            previous = latest_any_request(conn, item["run_num"])
            cursor = conn.execute(
                """
                INSERT INTO requests (
                    run_num, action, start_date, end_date, maker, input_date,
                    checker, approval_date, remark, pod, status, created_by,
                    created_at, updated_at, previous_request_id
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'approved', ?, ?, ?, ?)
                """,
                (
                    item["run_num"],
                    item["action"],
                    item["start_date"],
                    item["end_date"],
                    item["maker"],
                    item["input_date"],
                    item["checker"],
                    item["approval_date"],
                    item["remark"],
                    item["pod"],
                    item["maker"] or "import_history",
                    now,
                    now,
                    previous["id"] if previous else None,
                ),
            )
            imported_ids.append(cursor.lastrowid)

    return {
        "imported_count": len(imported_ids),
        "skipped_blank_run_num": skipped_blank_run_num,
        "imported_ids": imported_ids,
    }


def decide_request(request_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    decision = require_text(payload, "decision", "Decision").lower()
    if decision not in {"approve", "reject"}:
        raise ValueError("Decision must be approve or reject")
    checker = require_text(payload, "checker", "Checker")
    reject_reason = str(payload.get("reject_reason", "")).strip() or None

    status = "approved" if decision == "approve" else "rejected"
    now = datetime.now().isoformat(timespec="seconds")
    with connect() as conn:
        existing = conn.execute(
            "SELECT * FROM requests WHERE id = ?", (request_id,)
        ).fetchone()
        if existing is None:
            raise ValueError("Request not found")
        if existing["status"] != "pending":
            raise ValueError("Only pending requests can be approved or rejected")
        assigned_checker = (existing["checker"] or "").strip()
        if assigned_checker and assigned_checker.lower() != checker.lower():
            raise ValueError("This request is assigned to another checker")
        if not assigned_checker:
            raise ValueError("This request has no assigned checker")
        conn.execute(
            """
            UPDATE requests
            SET status = ?,
                checker = ?,
                approval_date = ?,
                reject_reason = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (status, checker, today().isoformat(), reject_reason, now, request_id),
        )
        row = conn.execute(
            "SELECT * FROM requests WHERE id = ?", (request_id,)
        ).fetchone()
        return row_to_dict(row)


def effective_export_rows(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    rows = conn.execute(
        """
        SELECT r.*
        FROM requests r
        JOIN (
            SELECT run_num, MAX(id) AS max_id
            FROM requests
            WHERE status IN ('approved', 'exported')
            GROUP BY run_num
        ) latest ON latest.max_id = r.id
        ORDER BY CAST(r.run_num AS INTEGER), r.run_num
        """,
    ).fetchall()
    return rows


def export_excel() -> dict[str, Any]:
    batch_id = datetime.now().strftime("export-%Y%m%d-%H%M%S")
    output_path = EXPORT_DIR / f"skip_loading_{batch_id}.xlsx"

    with connect() as conn:
        rows = effective_export_rows(conn)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Skip_Loading"
        headers = [
            "Run_Num",
            "Start_Date",
            "End_Date",
            "Maker",
            "Input_Date",
            "Checker",
            "Approval_Date",
            "REMARK",
            "POD",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

        for row in rows:
            sheet.append(
                [
                    row["run_num"],
                    excel_date_value(row["start_date"], "Start_Date"),
                    excel_date_value(row["end_date"], "End_Date"),
                    row["maker"],
                    excel_date_value(row["input_date"], "Input_Date"),
                    row["checker"] or "",
                    excel_date_value(row["approval_date"], "Approval_Date"),
                    row["remark"],
                    row["pod"],
                ]
            )

        for column in ("B", "C", "E", "G"):
            for cell in sheet[column][1:]:
                if cell.value:
                    cell.number_format = EXCEL_DATE_FORMAT

        widths = {
            "A": 12,
            "B": 14,
            "C": 14,
            "D": 18,
            "E": 14,
            "F": 20,
            "G": 16,
            "H": 44,
            "I": 12,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        if rows:
            table_ref = f"A1:I{len(rows) + 1}"
            table = Table(displayName="SkipLoadingTable", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)

        workbook.save(output_path)

        if rows:
            now = datetime.now().isoformat(timespec="seconds")
            row_ids = [row["id"] for row in rows]
            placeholders = ",".join("?" for _ in row_ids)
            conn.execute(
                f"""
                UPDATE requests
                SET status = 'exported',
                    export_batch_id = COALESCE(export_batch_id, ?),
                    updated_at = ?
                WHERE id IN ({placeholders})
                """,
                (batch_id, now, *row_ids),
            )

    return {
        "batch_id": batch_id,
        "file": str(output_path),
        "record_count": len(rows),
    }


def summary() -> dict[str, Any]:
    with connect() as conn:
        pending = conn.execute(
            "SELECT COUNT(*) AS count FROM requests WHERE status = 'pending'"
        ).fetchone()["count"]
        export_rows = effective_export_rows(conn)
        current_stop = sum(1 for row in export_rows if row["action"] == "STOP")
        current_resume = sum(1 for row in export_rows if row["action"] == "RESUME")
        return {
            "pending": pending,
            "will_export_now": len(export_rows),
            "current_stop": current_stop,
            "current_resume": current_resume,
            "today": today().isoformat(),
        }


def api_response(handler: SimpleHTTPRequestHandler, status: int, payload: Any) -> None:
    body = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def read_json(handler: SimpleHTTPRequestHandler) -> dict[str, Any]:
    length = int(handler.headers.get("Content-Length", "0"))
    if length == 0:
        return {}
    raw = handler.rfile.read(length)
    return json.loads(raw.decode("utf-8"))


def read_multipart_file(handler: SimpleHTTPRequestHandler, field_name: str) -> bytes:
    content_type = handler.headers.get("Content-Type", "")
    if "multipart/form-data" not in content_type:
        raise ValueError("Expected multipart/form-data upload")
    boundary_marker = "boundary="
    if boundary_marker not in content_type:
        raise ValueError("Upload boundary missing")
    boundary = content_type.split(boundary_marker, 1)[1].strip().strip('"')
    length = int(handler.headers.get("Content-Length", "0"))
    body = handler.rfile.read(length)
    delimiter = f"--{boundary}".encode()
    for part in body.split(delimiter):
        if f'name="{field_name}"'.encode() not in part:
            continue
        if b"\r\n\r\n" not in part:
            continue
        content = part.split(b"\r\n\r\n", 1)[1]
        content = content.rsplit(b"\r\n", 1)[0]
        if not content:
            raise ValueError("Uploaded file is empty")
        return content
    raise ValueError("Uploaded file field not found")


class AppHandler(SimpleHTTPRequestHandler):
    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/summary":
                api_response(self, 200, summary())
            elif parsed.path == "/api/requests":
                query = parse_qs(parsed.query)
                api_response(
                    self,
                    200,
                    list_requests(
                        status=query.get("status", [None])[0],
                        run_num=query.get("run_num", [None])[0],
                        checker=query.get("checker", [None])[0],
                        pod=query.get("pod", [None])[0],
                    ),
                )
            elif parsed.path == "/api/history":
                query = parse_qs(parsed.query)
                api_response(self, 200, get_history(query.get("run_num", [""])[0]))
            elif parsed.path == "/api/defaults":
                query = parse_qs(parsed.query)
                api_response(
                    self,
                    200,
                    get_defaults(
                        run_num=query.get("run_num", [None])[0],
                        action=query.get("action", [None])[0],
                    ),
                )
            elif parsed.path == "/api/export-preview":
                with connect() as conn:
                    rows = effective_export_rows(conn)
                    api_response(self, 200, [row_to_dict(row) for row in rows])
            elif parsed.path.startswith("/exports/"):
                self.serve_file(BASE_DIR / parsed.path.lstrip("/"))
            elif parsed.path == "/" or parsed.path.startswith("/static/"):
                if parsed.path == "/":
                    path = STATIC_DIR / "index.html"
                else:
                    path = STATIC_DIR / parsed.path.removeprefix("/static/")
                self.serve_file(path)
            else:
                api_response(self, 404, {"error": "Not found"})
        except Exception as exc:
            api_response(self, 400, {"error": str(exc)})

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/import-history":
                file_bytes = read_multipart_file(self, "file")
                api_response(self, 200, import_history_excel(file_bytes))
                return
            payload = read_json(self)
            if parsed.path == "/api/requests":
                api_response(self, 201, create_request(payload))
            elif parsed.path.startswith("/api/requests/") and parsed.path.endswith("/decision"):
                request_id = int(parsed.path.split("/")[3])
                api_response(self, 200, decide_request(request_id, payload))
            elif parsed.path == "/api/export":
                api_response(self, 200, export_excel())
            else:
                api_response(self, 404, {"error": "Not found"})
        except Exception as exc:
            api_response(self, 400, {"error": str(exc)})

    def serve_file(self, path: Path) -> None:
        if not path.exists() or not path.is_file():
            api_response(self, 404, {"error": "File not found"})
            return
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, format: str, *args: Any) -> None:
        print(f"{self.address_string()} - {format % args}")


def run() -> None:
    init_db()
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "8000"))
    server = ThreadingHTTPServer((host, port), AppHandler)
    print(f"Skip Loading Interface running at http://{host}:{port}")
    server.serve_forever()


if __name__ == "__main__":
    run()
